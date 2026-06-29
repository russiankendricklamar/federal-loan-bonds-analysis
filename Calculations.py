import requests
import pandas as pd
import numpy as np
from datetime import date, datetime, timedelta
from io import StringIO
from scipy.optimize import curve_fit
import warnings
import openpyxl
warnings.filterwarnings("ignore")

BASE_ISS = "https://iss.moex.com/iss"

def resolve_secid(isin: str) -> str | None:

    try:
        r = requests.get(f"{BASE_ISS}/securities.json",
                         params={"q": isin, "limit": 3}, timeout=15)
        r.raise_for_status()
        df = pd.DataFrame(r.json()["securities"]["data"],
                          columns=r.json()["securities"]["columns"])
        if not df.empty:
            return df.iloc[0]["secid"]
        print(f"[WARN] SECID не найден для {isin}")
        return None
    except Exception as e:
        print(f"[ERROR] {isin}: {e}")
        return None

BOARDS = [
    ("TQOB", "bonds"),
    ("TQCB", "bonds"),
    ("TQRD", "bonds"),
    ("TQEU", "bonds"),
]

FIELDS_SEC = [
    "SECID", "SHORTNAME", "ISIN", "FACEVALUE", "MATDATE",
    "COUPONVALUE", "COUPONPERCENT", "COUPONPERIOD", "ACCRUEDINT",
    "NEXTCOUPON", "PREVLEGALCLOSEPRICE", "CURRENCYID",
]
FIELDS_MKT = [
    "SECID", "LAST", "WAPRICE", "LCLOSEPRICE", "YIELD",
    "DURATION", "ZSPREAD",
]


def get_ofz_marketdata(secids: list[str]) -> pd.DataFrame:
    """
    Ищет бумаги по нескольким торговым доскам ISS MOEX.
    Возвращает DataFrame с индексом SECID.
    MATDATE приведён к datetime, числовые поля — к float.
    """
    frames    = []
    remaining = set(secids)

    for board, market in BOARDS:
        if not remaining:
            break

        url = (f"{BASE_ISS}/engines/stock/markets/{market}/"
               f"boards/{board}/securities.json")
        params = {
            "securities":    ",".join(remaining),
            "iss.meta":      "off",
            "iss.only":      "securities,marketdata",
            "securities.columns": ",".join(FIELDS_SEC),
            "marketdata.columns": ",".join(FIELDS_MKT),
        }

        try:
            r    = requests.get(url, params=params, timeout=15)
            r.raise_for_status()
            data = r.json()
        except Exception as e:
            print(f"  [{board}] ошибка запроса: {e}")
            continue

        sec = pd.DataFrame(data["securities"]["data"],
                           columns=data["securities"]["columns"])
        mkt = pd.DataFrame(data["marketdata"]["data"],
                           columns=data["marketdata"]["columns"])

        if sec.empty:
            continue

        merged = sec.merge(mkt, on="SECID", how="left").set_index("SECID")

        merged["MATDATE"]    = pd.to_datetime(merged["MATDATE"],    errors="coerce")
        merged["NEXTCOUPON"] = pd.to_datetime(merged["NEXTCOUPON"], errors="coerce")
        for col in ["FACEVALUE", "ACCRUEDINT", "COUPONVALUE", "COUPONPERCENT",
                    "COUPONPERIOD", "LAST", "WAPRICE", "LCLOSEPRICE",
                    "YIELD", "DURATION", "ZSPREAD", "PREVLEGALCLOSEPRICE"]:
            if col in merged.columns:
                merged[col] = pd.to_numeric(merged[col], errors="coerce")

        frames.append(merged)
        found      = set(merged.index.tolist())
        remaining -= found
        print(f"  [{board}] найдено {len(found)}: {found}")

    if remaining:
        print(f"[WARN] Не найдены ни на одной доске: {remaining}")

    if not frames:
        return pd.DataFrame()

    result = pd.concat(frames)
    result = result[~result.index.duplicated(keep="first")]
    return result

def get_history(secid: str, from_date: str, till_date: str) -> pd.DataFrame:

    url = (f"{BASE_ISS}/history/engines/stock/markets/bonds"
           f"/boards/TQOB/securities/{secid}.json")
    params = {
        "from": from_date,
        "till": till_date,
        "limit": 500,
        "history.columns": "TRADEDATE,CLOSE,YIELDATWAP,VOLUME,WAPRICE",
    }
    r = requests.get(url, params=params, timeout=30)
    r.raise_for_status()
    data = r.json()
    df = pd.DataFrame(data["history"]["data"],
                      columns=data["history"]["columns"])
    df["TRADEDATE"] = pd.to_datetime(df["TRADEDATE"])
    df = df.set_index("TRADEDATE").apply(pd.to_numeric, errors="coerce")
    return df.dropna(subset=["CLOSE"])

def get_entry_price(secid: str, entry_date: str) -> float:

    url = (f"{BASE_ISS}/history/engines/stock/markets/bonds"
           f"/boards/TQOB/securities/{secid}.json")
    params = {
        "from": entry_date,
        "till": entry_date,
        "history.columns": "TRADEDATE,CLOSE,WAPRICE,FACEVALUE",
    }
    r = requests.get(url, params=params, timeout=15)
    r.raise_for_status()
    data = r.json()
    rows = data["history"]["data"]
    if not rows:
        print(f"[WARN] {secid}: нет торгов на {entry_date}, используем текущую цену")
        return np.nan
    row      = dict(zip(data["history"]["columns"], rows[0]))
    face     = float(row["FACEVALUE"]) if row["FACEVALUE"] else 1000.0
    waprice  = float(row["WAPRICE"]) if row["WAPRICE"] else None
    close    = float(row["CLOSE"])   if row["CLOSE"]   else None
    pct      = waprice or close
    if pct is None:
        print(f"[WARN] {secid}: нет цены на {entry_date}")
        return np.nan
    return round(pct / 100 * face, 4)  # рубли

def get_moex_zcyc() -> dict:

    r = requests.get("https://iss.moex.com/iss/engines/stock/zcyc.json", timeout=15)
    r.raise_for_status()
    data = r.json()

    params_row = data["params"]["data"][0]
    params_cols = data["params"]["columns"]
    params = dict(zip(params_cols, params_row))

    yy = pd.DataFrame(data["yearyields"]["data"],
                      columns=data["yearyields"]["columns"])
    yy = yy.set_index("period")["value"].to_dict()

    tradedate = params["tradedate"]
    tradetime = params["tradetime"]
    print(f"G-Curve MOEX загружена: {tradedate} {tradetime}, "
          f"тенора: {list(yy.keys())}")
    return {"params": params, "yearyields": yy,
            "tradedate": tradedate, "tradetime": tradetime}

def gcurve_rate(tau: float, zcyc: dict) -> float:

    from scipy.interpolate import CubicSpline
    yy = zcyc["yearyields"]
    tenors = sorted(yy.keys())
    rates  = [yy[t] for t in tenors]
    if tau <= tenors[0]:
        return rates[0]
    if tau >= tenors[-1]:
        return rates[-1]
    cs = CubicSpline(tenors, rates)
    return float(cs(tau))

def z_spread_gcurve(ytm_pct: float, mod_dur: float, zcyc: dict) -> float:

    kbd = gcurve_rate(mod_dur, zcyc)
    return round((ytm_pct - kbd) * 100, 2)  # в bps

def get_cashflows(secid: str, qty: int, face: float) -> pd.DataFrame:

    url    = f"{BASE_ISS}/securities/{secid}/bondization.json"
    params = {"iss.meta": "off", "iss.only": "coupons,amortizations",
              "limit": "unlimited"}
    try:
        r    = requests.get(url, params=params, timeout=15)
        data = r.json()
    except Exception as e:
        print(f"  [CF] {secid} ошибка: {e}")
        return pd.DataFrame()

    rows = []

    coup_cols = data["coupons"]["columns"]
    for rec in data["coupons"]["data"]:
        d = dict(zip(coup_cols, rec))
        dt = pd.to_datetime(d.get("coupondate"), errors="coerce")
        if pd.isnull(dt) or dt.date() < date.today():
            continue
        val = d.get("value")
        if val is None:
            prc = d.get("valueprc")
            val = float(prc) / 100 * face if prc else 0.0
        rows.append({
            "date":       dt,
            "type":       "coupon",
            "amount_rub": float(val) * qty,
        })

    amort_cols = data["amortizations"]["columns"]
    for rec in data["amortizations"]["data"]:
        d  = dict(zip(amort_cols, rec))
        dt = pd.to_datetime(d.get("amortdate"), errors="coerce")
        if pd.isnull(dt) or dt.date() < date.today():
            continue
        val = d.get("value")
        if val is None:
            prc = d.get("valueprc")
            val = float(prc) / 100 * face if prc else 0.0
        rows.append({
            "date":       dt,
            "type":       "amortization",
            "amount_rub": float(val) * qty,
        })

    return pd.DataFrame(rows) if rows else pd.DataFrame()

def build_cf_ladder(analytics: pd.DataFrame,
                    positions: list[dict],
                    freq: str = "ME") -> tuple[pd.DataFrame, pd.DataFrame]:

    all_cf = []

    for pos in positions:
        secid = pos["secid"]
        qty   = pos["qty"]
        face_row = analytics[analytics["SECID"] == secid]
        face     = float(face_row["Face"].iloc[0]) if not face_row.empty else 1000.0

        cf = get_cashflows(secid, qty, face)
        if cf.empty:
            print(f"  [CF] {secid}: нет данных")
            continue
        cf["secid"] = secid
        all_cf.append(cf)
        print(f"  [CF] {secid}: {len(cf)} событий, "
              f"сумма {cf['amount_rub'].sum():,.0f} руб.")

    if not all_cf:
        print("[WARN] Cash flow ladder пустой")
        return pd.DataFrame(), pd.DataFrame()

    df = pd.concat(all_cf, ignore_index=True)
    df["period"] = df["date"].dt.to_period(freq[0])

    ladder_agg = (df.groupby(["period", "type"])["amount_rub"]
                    .sum()
                    .unstack(fill_value=0)
                    .reset_index())
    for col in ["coupon", "amortization"]:
        if col not in ladder_agg.columns:
            ladder_agg[col] = 0.0
    ladder_agg["total"] = ladder_agg["coupon"] + ladder_agg["amortization"]
    ladder_agg = ladder_agg.sort_values("period").reset_index(drop=True)
    ladder_agg["period_str"] = ladder_agg["period"].astype(str)

    ladder_wide = (df.groupby(["period", "secid"])["amount_rub"]
                     .sum()
                     .unstack(fill_value=0)
                     .reset_index())
    ladder_wide["period_str"] = ladder_wide["period"].astype(str)
    ladder_wide = ladder_wide.sort_values("period").reset_index(drop=True)

    return ladder_wide, ladder_agg

class OFZPortfolio:

    FREQ = 2

    def __init__(self, positions: list[dict], settlement: date = None):
        self.positions  = positions
        self.settlement = settlement or (date.today() + timedelta(days=1))
        self._analytics: pd.DataFrame | None = None

    def compute(self, zcyc: dict = None) -> pd.DataFrame:

        secids = [p["secid"] for p in self.positions]
        mkt_df = get_ofz_marketdata(secids)
        rows   = []

        for pos in self.positions:
            secid = pos["secid"]
            qty   = pos["qty"]
            cost  = pos["cost_price"]

            if secid not in mkt_df.index:
                print(f"[SKIP] {secid} нет в маркетдате TQOB")
                continue

            row  = mkt_df.loc[secid]
            face = row["FACEVALUE"]  if not np.isnan(row["FACEVALUE"])  else 1000.0
            ai   = row["ACCRUEDINT"] if not np.isnan(row["ACCRUEDINT"]) else 0.0

            last_pct = row["LAST"]
            if np.isnan(last_pct): last_pct = row["WAPRICE"]
            if np.isnan(last_pct): last_pct = row["LCLOSEPRICE"]
            if np.isnan(last_pct):
                last_rub = cost if cost is not None else np.nan
                last_pct = np.nan
                print(f"[WARN] {secid}: нет рыночной цены, используем cost_price")
            else:
                last_rub = last_pct / 100 * face

            ytm_pct       = row["YIELD"]
            mod_dur_years = row["DURATION"] / 365.0 \
                            if not np.isnan(row["DURATION"]) else np.nan

            mac_dur = mod_dur_years * (1 + ytm_pct / (100 * self.FREQ)) \
                      if not np.isnan(mod_dur_years) and not np.isnan(ytm_pct) \
                      else np.nan

            convex_approx = mod_dur_years ** 2 + mod_dur_years \
                            if not np.isnan(mod_dur_years) else np.nan

            if zcyc is not None and not np.isnan(ytm_pct) and not np.isnan(mod_dur_years):
                zs_bps = z_spread_gcurve(ytm_pct, mod_dur_years, zcyc)
            else:
                zs_raw = row["ZSPREAD"]
                zs_bps = zs_raw * 100 if not np.isnan(zs_raw) else np.nan

            dirty_mkt = last_rub + ai if not np.isnan(last_rub) else np.nan
            mkt_value = dirty_mkt * qty if not np.isnan(dirty_mkt) else np.nan
            pnl       = (last_rub - cost) * qty \
                        if (cost is not None and not np.isnan(last_rub)) else np.nan

            dv01 = mod_dur_years * dirty_mkt * 0.0001 * qty \
                   if not np.isnan(mod_dur_years) and not np.isnan(dirty_mkt) \
                   else np.nan

            def _r(x, n=4):
                return round(x, n) if (x is not None and not np.isnan(x)) else np.nan

            rows.append({
                "SECID":         secid,
                "Name":          row["SHORTNAME"],
                "Maturity":      row["MATDATE"].date() if pd.notnull(row["MATDATE"]) else None,
                "Qty":           qty,
                "CostPrice":     _r(cost, 2),
                "LastPrice_rub": _r(last_rub, 2),
                "LastPrice_pct": _r(last_pct, 4),
                "AI":            _r(ai, 2),
                "Face":          face,
                "DirtyMkt":      _r(dirty_mkt, 2),
                "MktValue":      _r(mkt_value, 2),
                "YTM_%":         _r(ytm_pct, 4),
                "MacDur_years":  _r(mac_dur, 4),
                "ModDur_years":  _r(mod_dur_years, 4),
                "Convex_approx": _r(convex_approx, 4),
                "DV01":          _r(dv01, 2),
                "ZSpread_bps":   _r(zs_bps, 2),
                "PnL":           _r(pnl, 2),
            })

        self._analytics = pd.DataFrame(rows)
        return self._analytics

    def summary(self) -> dict:

        if self._analytics is None:
            raise RuntimeError("Сначала вызовите .compute()")
        df = self._analytics.dropna(subset=["MktValue"])
        if df.empty:
            print("[WARN] Аналитика пустая")
            return {}

        total_mv = df["MktValue"].sum()
        w        = df["MktValue"] / total_mv

        port_modd  = (df["ModDur_years"] * w).sum()
        dollar_dur = port_modd * total_mv

        return {
            "TotalMktValue":   round(total_mv, 2),
            "TotalPnL":        round(df["PnL"].dropna().sum(), 2),
            "PortfolioDV01":   round(df["DV01"].sum(), 2),

            "PortMacaulayDur": round((df["MacDur_years"] * w).sum(), 4),
            "PortModDur":      round(port_modd, 4),
            "PortConvexity":   round((df["Convex_approx"] * w).sum(), 4),
            "DollarDuration":  round(dollar_dur, 2),

            "WeightedYTM_%":   round((df["YTM_%"] * w).sum(), 4),
            "WeightedZSpread": round((df["ZSpread_bps"] * w).sum(), 4),
        }

    def scenario_shift(self, shift_bps: float) -> pd.DataFrame:

        if self._analytics is None:
            raise RuntimeError("Сначала вызовите .compute()")
        df = self._analytics.dropna(subset=["ModDur_years", "DirtyMkt"]).copy()
        if df.empty:
            return df

        shift = shift_bps / 10000

        df["dP"] = (
            -df["ModDur_years"] * shift
            + 0.5 * df["Convex_approx"] * shift ** 2
        ) * df["DirtyMkt"] * df["Qty"]

        df["ModDur_stressed"] = (df["ModDur_years"]
                                 - df["Convex_approx"] * shift).round(4)
        df["dModDur"]         = (df["ModDur_stressed"]
                                 - df["ModDur_years"]).round(4)

        ytm_stressed          = df["YTM_%"] + shift * 100
        df["MacDur_stressed"] = (df["ModDur_stressed"]
                                 * (1 + ytm_stressed / (100 * self.FREQ))).round(4)
        df["dMacDur"]         = (df["MacDur_stressed"]
                                 - df["MacDur_years"]).round(4)

        dirty_stressed      = df["DirtyMkt"] + df["dP"] / df["Qty"]
        df["DV01_stressed"] = (df["ModDur_stressed"]
                               * dirty_stressed * 0.0001 * df["Qty"]).round(2)
        df["dDV01"]         = (df["DV01_stressed"] - df["DV01"]).round(2)

        return df[[
            "SECID",        "Name",
            "MktValue",     "dP",
            "MacDur_years", "MacDur_stressed", "dMacDur",
            "ModDur_years", "ModDur_stressed", "dModDur",
            "DV01",         "DV01_stressed",   "dDV01",
        ]]

    def pnl_attribution(self) -> pd.DataFrame:

        if self._analytics is None:
            raise RuntimeError("Сначала вызовите .compute()")
        df = self._analytics.copy()
        df["Carry_PnL"] = df["AI"] * df["Qty"]
        return df[["SECID", "Name", "PnL", "Carry_PnL"]]

import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage

HDR_FILL  = PatternFill("solid", fgColor="1F4E79")
HDR_FONT  = Font(bold=True, color="FFFFFF", size=10)
SUB_FILL  = PatternFill("solid", fgColor="2E75B6")
ALT_FILL  = PatternFill("solid", fgColor="DEEAF1")
FX_FILL   = PatternFill("solid", fgColor="FFFACD")
FX_NOTE   = PatternFill("solid", fgColor="FFE699")
FX_FONT   = Font(size=10, color="7F6000", italic=True)
BODY_FONT = Font(size=10)
CENTER    = Alignment(horizontal="center", vertical="center")
LEFT      = Alignment(horizontal="left",   vertical="center")
thin      = Side(style="thin", color="B0C4DE")
BORDER    = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_header(ws, row_idx: int, fill=HDR_FILL, font=HDR_FONT):
    for cell in ws[row_idx]:
        if cell.value is not None:
            cell.fill = fill; cell.font = font
            cell.alignment = CENTER; cell.border = BORDER

def style_rows(ws, start: int, end: int):
    for i, row in enumerate(ws.iter_rows(min_row=start, max_row=end)):
        bg = ALT_FILL if i % 2 == 0 else PatternFill()
        for cell in row:
            cell.font = BODY_FONT; cell.fill = bg
            cell.border = BORDER; cell.alignment = CENTER

def autofit(ws):
    for col in ws.columns:
        valid = [c for c in col
                 if hasattr(c, "column_letter") and c.value is not None]
        if not valid:
            continue
        max_len = max((len(str(c.value)) for c in valid), default=8)
        ws.column_dimensions[valid[0].column_letter].width = min(max_len + 3, 30)

def write_df(ws, df: pd.DataFrame, start_row: int = 1) -> int:
    for r_idx, row in enumerate(
            dataframe_to_rows(df, index=False, header=True), start_row):
        for c_idx, val in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=val)
    style_header(ws, start_row)
    style_rows(ws, start_row + 1, start_row + len(df))
    autofit(ws)
    return start_row + len(df) + 1

def plot_cf_ladder(ladder_agg: pd.DataFrame,
                   ladder_wide: pd.DataFrame,
                   freq_label: str = "месяц") -> None:
    import matplotlib.pyplot as plt
    import matplotlib.gridspec as gridspec
    import matplotlib.ticker as mticker

    if ladder_agg.empty:
        print("[WARN] Нет данных для CF графика")
        return

    periods = ladder_agg["period_str"].tolist()
    x       = np.arange(len(periods))

    fig = plt.figure(figsize=(20, 16))
    gs  = gridspec.GridSpec(3, 1, figure=fig,
                            hspace=0.5, height_ratios=[3, 2, 2])

    ax1 = fig.add_subplot(gs[0])
    ax1.bar(x, ladder_agg["coupon"],
            color="#2196F3", alpha=0.85, label="Купоны", zorder=3)
    ax1.bar(x, ladder_agg["amortization"],
            color="#FF9800", alpha=0.85,
            bottom=ladder_agg["coupon"],
            label="Погашение / Амортизация", zorder=3)
    for i, (c, a) in enumerate(zip(ladder_agg["coupon"],
                                   ladder_agg["amortization"])):
        total = c + a
        if total > 0:
            ax1.text(i, total + total * 0.01,
                     f"{total/1e6:.1f}M" if total >= 1e6 else f"{total:,.0f}",
                     ha="center", va="bottom", fontsize=7, rotation=45)
    ax1.set_xticks(x)
    ax1.set_xticklabels(periods, rotation=45, ha="right", fontsize=8)
    ax1.yaxis.set_major_formatter(
        mticker.FuncFormatter(lambda v, _: f"{v/1e6:.1f}M" if v >= 1e6
                              else f"{v:,.0f}"))
    ax1.set_ylabel("Денежный поток (руб.)"); ax1.set_title(f"Cash Flow Ladder — по {freq_label}ам")
    ax1.legend(); ax1.grid(True, alpha=0.3, axis="y", zorder=0)

    ax2       = fig.add_subplot(gs[1])
    cum_total = ladder_agg["total"].cumsum()
    cum_coup  = ladder_agg["coupon"].cumsum()
    cum_amort = ladder_agg["amortization"].cumsum()
    ax2.fill_between(x, cum_total, alpha=0.15, color="#4CAF50")
    ax2.plot(x, cum_total, lw=2,   color="#4CAF50", label="Итого накопленно")
    ax2.plot(x, cum_coup,  lw=1.5, color="#2196F3", ls="--", label="Купоны накопленно")
    ax2.plot(x, cum_amort, lw=1.5, color="#FF9800", ls="--", label="Погашения накопленно")
    ax2.set_xticks(x)
    ax2.set_xticklabels(periods, rotation=45, ha="right", fontsize=8)
    ax2.yaxis.set_major_formatter(
        mticker.FuncFormatter(lambda v, _: f"{v/1e6:.1f}M" if v >= 1e6
                              else f"{v:,.0f}"))
    ax2.set_ylabel("Накопленный CF (руб.)"); ax2.set_title("Накопленный денежный поток")
    ax2.legend(fontsize=9); ax2.grid(True, alpha=0.3, zorder=0)

    ax3    = fig.add_subplot(gs[2])
    secids = [c for c in ladder_wide.columns
              if c not in ("period", "period_str")]
    if secids:
        heat_df          = ladder_wide[secids].T
        heat_df.columns  = periods
        im = ax3.imshow(heat_df.values, aspect="auto", cmap="YlOrRd",
                        interpolation="nearest")
        ax3.set_xticks(range(len(periods)))
        ax3.set_xticklabels(periods, rotation=45, ha="right", fontsize=7)
        ax3.set_yticks(range(len(secids)))
        ax3.set_yticklabels([s[-7:] for s in secids], fontsize=9)
        ax3.set_title("Тепловая карта денежных потоков по бумагам")
        plt.colorbar(im, ax=ax3,
                     format=mticker.FuncFormatter(
                         lambda v, _: f"{v/1e6:.1f}M" if v >= 1e6
                         else f"{v:,.0f}"))

    plt.suptitle(f"Cash Flow Ladder — {date.today()}", fontsize=14,
                 fontweight="bold", y=1.01)
    plt.savefig("cf_ladder.png", dpi=150, bbox_inches="tight")
    plt.close()
    print("Сохранён cf_ladder.png")

def breakeven_analysis(analytics: pd.DataFrame,
                       positions: list[dict],
                       horizon_days: int = 365) -> pd.DataFrame:

    rows = []
    ann_df = analytics.set_index("SECID")

    for pos in positions:
        secid = pos["secid"]
        qty   = pos["qty"]

        if secid not in ann_df.index:
            continue

        row      = ann_df.loc[secid]
        mkt_val  = float(row.get("MktValue",  0) or 0)
        mod_dur  = float(row.get("ModDur_years", 0) or 0)
        coup_pct = float(row.get("CouponPct",  0) or 0)   # % годовых
        face     = float(row.get("Face",       1000) or 1000)
        name     = str(row.get("Name", secid))

        if mkt_val == 0 or mod_dur == 0:
            continue

        carry_rub = (coup_pct / 100) * face * qty * (horizon_days / 365)

        dv01 = mod_dur * mkt_val / 10_000

        be_shift = carry_rub / dv01 if dv01 > 0 else float("inf")

        be_shift_ann = be_shift * (365 / horizon_days)

        def margin(shift_bps: float) -> float:
            dp = -mod_dur * (shift_bps / 10_000) * mkt_val
            return carry_rub + dp

        rows.append({
            "SECID":           secid,
            "Name":            name,
            "MktValue":        round(mkt_val,       2),
            "CouponPct":       round(coup_pct,       4),
            "ModDur":          round(mod_dur,        4),
            "DV01":            round(dv01,            2),
            f"Carry_{horizon_days}d_rub": round(carry_rub, 2),
            "BE_shift_bps":    round(be_shift,        1),
            "BE_shift_ann_bps":round(be_shift_ann,    1),
            "Margin_+100bps":  round(margin(100),     2),
            "Margin_+200bps":  round(margin(200),     2),
            "Margin_+300bps":  round(margin(300),     2),
        })

    df = pd.DataFrame(rows)
    if df.empty:
        return df

    port_carry   = df[f"Carry_{horizon_days}d_rub"].sum()
    port_dv01    = df["DV01"].sum()
    port_mkt     = df["MktValue"].sum()
    port_be      = port_carry / port_dv01 if port_dv01 > 0 else float("inf")
    port_be_ann  = port_be * (365 / horizon_days)
    port_mod_dur = (df["ModDur"] * df["MktValue"]).sum() / port_mkt

    total = {
        "SECID":           "◆ ПОРТФЕЛЬ",
        "Name":            "",
        "MktValue":        round(port_mkt,  2),
        "CouponPct":       "",
        "ModDur":          round(port_mod_dur, 4),
        "DV01":            round(port_dv01, 2),
        f"Carry_{horizon_days}d_rub": round(port_carry, 2),
        "BE_shift_bps":    round(port_be,   1),
        "BE_shift_ann_bps":round(port_be_ann, 1),
        "Margin_+100bps":  round(port_carry - 100 * port_dv01, 2),
        "Margin_+200bps":  round(port_carry - 200 * port_dv01, 2),
        "Margin_+300bps":  round(port_carry - 300 * port_dv01, 2),
    }
    df = pd.concat([df, pd.DataFrame([total])], ignore_index=True)
    return df

def plot_breakeven(be_df: pd.DataFrame, horizon_days: int = 365) -> None:

    import matplotlib.pyplot as plt
    import matplotlib.ticker as mticker

    pos_df   = be_df[be_df["SECID"] != "◆ ПОРТФЕЛЬ"].copy()
    port_row = be_df[be_df["SECID"] == "◆ ПОРТФЕЛЬ"].iloc[0]
    carry_col = [c for c in be_df.columns if c.startswith("Carry_")][0]

    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(18, 7))
    fig.suptitle(f"Break-Even Analysis — горизонт {horizon_days} дней "
                 f"({date.today()})", fontsize=13, fontweight="bold")

    labels   = pos_df["SECID"].str[-7:].tolist() + ["ПОРТФЕЛЬ"]
    be_vals  = pos_df["BE_shift_bps"].tolist() + [float(port_row["BE_shift_bps"])]
    colors   = ["#2196F3"] * len(pos_df) + ["#FF5722"]

    x    = range(len(labels))
    bars = ax1.bar(x, be_vals, color=colors, alpha=0.85, width=0.6)

    for shift, color, ls in [(100, "#FF9800", "--"),
                              (200, "#F44336", "-."),
                              (300, "#B71C1C", ":")]:
        ax1.axhline(shift, color=color, lw=1.5, ls=ls,
                    label=f"{shift} bps")

    for bar, val in zip(bars, be_vals):
        ax1.text(bar.get_x() + bar.get_width() / 2,
                 val + max(be_vals) * 0.01,
                 f"{val:.0f}", ha="center", va="bottom", fontsize=9,
                 fontweight="bold" if val == float(port_row["BE_shift_bps"]) else "normal")

    ax1.set_xticks(list(x))
    ax1.set_xticklabels(labels, rotation=20, ha="right", fontsize=9)
    ax1.set_ylabel("Break-Even сдвиг (bps)")
    ax1.set_title("При каком сдвиге carry = capital loss")
    ax1.legend(title="Стресс-уровни", fontsize=8)
    ax1.grid(True, alpha=0.3, axis="y")

    port_carry = float(port_row[carry_col])
    port_dv01  = float(port_row["DV01"])

    shifts     = range(0, 501, 10)
    cap_losses = [shift * port_dv01 for shift in shifts]
    carry_line = [port_carry] * len(shifts)
    net        = [port_carry - cl for cl in cap_losses]

    ax2.fill_between(shifts, net, 0,
                     where=[n >= 0 for n in net],
                     alpha=0.15, color="#4CAF50", label="Прибыль (carry > loss)")
    ax2.fill_between(shifts, net, 0,
                     where=[n < 0 for n in net],
                     alpha=0.15, color="#F44336", label="Убыток (loss > carry)")
    ax2.plot(shifts, carry_line, lw=2, color="#4CAF50",
             ls="--", label=f"Carry {port_carry:,.0f} руб.")
    ax2.plot(shifts, cap_losses, lw=2, color="#F44336",
             label="Capital Loss (ModDur × shift × MV)")
    ax2.plot(shifts, net,        lw=2, color="#2196F3",
             label="Net P&L (carry − loss)")

    be_x = float(port_row["BE_shift_bps"])
    ax2.axvline(be_x, color="black", lw=1.5, ls="--")
    ax2.text(be_x + 5, port_carry * 0.05,
             f"BE = {be_x:.0f} bps", fontsize=10, color="black",
             fontweight="bold")

    ax2.yaxis.set_major_formatter(
        mticker.FuncFormatter(lambda v, _: f"{v/1e6:.1f}M" if abs(v) >= 1e6
                              else f"{v:,.0f}"))
    ax2.set_xlabel("Параллельный сдвиг G-Curve (bps)")
    ax2.set_ylabel("Руб.")
    ax2.set_title("Carry vs Capital Loss — портфель")
    ax2.legend(fontsize=8)
    ax2.grid(True, alpha=0.3)
    ax2.axhline(0, color="black", lw=0.8)

    plt.tight_layout()
    plt.savefig("breakeven.png", dpi=150, bbox_inches="tight")
    plt.close()
    print("Сохранён breakeven.png")

KRD_TENORS = [0.25, 0.5, 1.0, 2.0, 3.0, 5.0, 7.0, 10.0, 15.0, 20.0]
KRD_SHIFT  = 0.0100   # 100 bps = 1%


def _reprice_with_tenor_bump(cf_rows: list[dict],
                              zcyc_base: dict,
                              tenor_key: float,
                              shift: float,
                              today) -> float:

    from math import exp

    yy      = zcyc_base["yearyields"]
    tenors  = sorted(yy.keys())

    lo_list = [t for t in tenors if t <  tenor_key]
    hi_list = [t for t in tenors if t >  tenor_key]
    t_lo    = lo_list[-1] if lo_list else tenor_key
    t_hi    = hi_list[0]  if hi_list else tenor_key

    def hat_weight(t: float) -> float:
        eps = 1e-9

        is_leftmost  = tenor_key <= tenors[0]  + eps
        is_rightmost = tenor_key >= tenors[-1] - eps

        if is_rightmost:
            if t >= tenor_key - eps:
                return 1.0
            elif t >= t_lo:
                span = tenor_key - t_lo
                return (t - t_lo) / span if span > eps else 1.0
            return 0.0

        if is_leftmost:
            if t <= tenor_key + eps:
                return 1.0
            elif t <= t_hi:
                span = t_hi - tenor_key
                return (t_hi - t) / span if span > eps else 1.0
            return 0.0

        if t <= t_lo + eps or t >= t_hi - eps:
            return 0.0
        span_lo = tenor_key - t_lo
        span_hi = t_hi - tenor_key
        if t <= tenor_key:
            return (t - t_lo) / span_lo if span_lo > eps else 1.0
        else:
            return (t_hi - t) / span_hi if span_hi > eps else 1.0

    def base_rate(t: float) -> float:

        if t <= tenors[0]:  return yy[tenors[0]]
        if t >= tenors[-1]: return yy[tenors[-1]]
        for i in range(len(tenors) - 1):
            if tenors[i] <= t <= tenors[i + 1]:
                w = (t - tenors[i]) / (tenors[i + 1] - tenors[i])
                return yy[tenors[i]] * (1 - w) + yy[tenors[i + 1]] * w
        return yy[tenors[-1]]

    pv = 0.0
    for cf in cf_rows:
        cf_date = cf.get("date")
        amount  = cf.get("amount", 0)
        if cf_date is None or amount <= 0:
            continue
        try:
            t = (pd.Timestamp(cf_date).date() - today).days / 365.25
        except Exception:
            continue
        if t <= 0:
            continue

        r_base   = base_rate(t) / 100.0
        w        = hat_weight(t)
        r_bumped = r_base + w * shift
        pv      += amount * exp(-r_bumped * t)

    return pv

def compute_krd(analytics: pd.DataFrame,
                positions: list[dict],
                zcyc: dict,
                cf_map: dict[str, list[dict]],
                tenors: list[float] = KRD_TENORS,
                shift_bps: float = 1.0) -> pd.DataFrame:
    from datetime import date as _date

    today = _date.today()
    shift = shift_bps / 10_000
    ann_df = analytics.set_index("SECID")
    rows = []

    for pos in positions:
        secid   = pos["secid"]
        cf_rows = cf_map.get(secid, [])

        if not cf_rows or secid not in ann_df.index:
            continue

        row = ann_df.loc[secid]

        p0 = float(row.get("MktValue", 0) or 0)
        if p0 == 0:
            print(f"  [WARN] {secid}: MktValue=0, пропускаем")
            continue

        krd_row = {
            "SECID":    secid,
            "Name":     str(row.get("Name", secid)),
            "MktValue": round(p0, 2),
        }

        krd_sum = 0.0
        for tenor in tenors:
            p_plus  = _reprice_with_tenor_bump(cf_rows, zcyc, tenor, +shift, today)
            p_minus = _reprice_with_tenor_bump(cf_rows, zcyc, tenor, -shift, today)

            krd = (p_minus - p_plus) / (2 * shift * p0)
            krd_row[f"KRD_{tenor}"] = round(krd, 4)
            krd_sum += krd

        krd_row["KRD_sum"] = round(krd_sum, 4)
        krd_row["ModDur"]  = round(float(row.get("ModDur_years", 0) or 0), 4)
        rows.append(krd_row)

    if not rows:
        return pd.DataFrame()

    df = pd.DataFrame(rows)

    total_mv     = df["MktValue"].sum()
    port_mod_dur = (df["ModDur"] * df["MktValue"]).sum() / total_mv

    total = {
        "SECID":    "◆ ПОРТФЕЛЬ",
        "Name":     "",
        "MktValue": round(total_mv, 2),
        "ModDur":   round(port_mod_dur, 4),
    }
    for tenor in tenors:
        col = f"KRD_{tenor}"
        if col in df.columns:
            total[col] = round(
                (df[col] * df["MktValue"]).sum() / total_mv, 4)
    total["KRD_sum"] = round(
        sum(total.get(f"KRD_{t}", 0) for t in tenors), 4)

    df = pd.concat([df, pd.DataFrame([total])], ignore_index=True)
    return df

def plot_krd(krd_df: pd.DataFrame, tenors: list[float] = KRD_TENORS) -> None:
    import matplotlib.pyplot as plt
    import matplotlib.ticker as mticker
    import numpy as np

    pos_df   = krd_df[krd_df["SECID"] != "◆ ПОРТФЕЛЬ"].copy()
    port_row = krd_df[krd_df["SECID"] == "◆ ПОРТФЕЛЬ"]

    fig, axes = plt.subplots(1, 2, figsize=(18, 7))
    fig.suptitle(f"Key Rate Duration (KRD) — {date.today()}",
                 fontsize=13, fontweight="bold")

    ax = axes[0]
    tenor_labels = [str(t) for t in tenors]
    x = np.arange(len(tenors))
    cmap = plt.cm.get_cmap("tab10", len(pos_df))

    for idx, (_, row) in enumerate(pos_df.iterrows()):
        krd_vals = [float(row.get(f"KRD_{t}", 0) or 0) for t in tenors]
        ax.plot(x, krd_vals, marker="o", lw=1.8,
                color=cmap(idx), label=row["SECID"][-7:], alpha=0.85)

    if not port_row.empty:
        port_krd = [float(port_row.iloc[0].get(f"KRD_{t}", 0) or 0) for t in tenors]
        ax.plot(x, port_krd, marker="D", lw=2.5,
                color="black", label="◆ ПОРТФЕЛЬ", zorder=5)

    ax.set_xticks(x)
    ax.set_xticklabels(tenor_labels)
    ax.set_xlabel("Тенор G-Curve (лет)")
    ax.set_ylabel("KRD (лет)")
    ax.set_title("KRD-профиль по тенорам")
    ax.legend(fontsize=8)
    ax.grid(True, alpha=0.3)
    ax.axhline(0, color="black", lw=0.7, ls="--")

    ax2 = axes[1]
    if not port_row.empty:
        port_krd = [float(port_row.iloc[0].get(f"KRD_{t}", 0) or 0) for t in tenors]
        bars = ax2.bar(tenor_labels, port_krd,
                       color=["#1a6fad" if v >= 0 else "#ad2a1a" for v in port_krd],
                       alpha=0.85, width=0.6)
        for bar, val in zip(bars, port_krd):
            ax2.text(bar.get_x() + bar.get_width() / 2,
                     val + max(port_krd) * 0.02,
                     f"{val:.3f}", ha="center", fontsize=9)

        krd_sum = float(port_row.iloc[0].get("KRD_sum", 0) or 0)
        mod_dur = float(port_row.iloc[0].get("ModDur", 0) or 0)
        ax2.axhline(krd_sum / len(tenors), color="green",  lw=1.5, ls="--",
                    label=f"KRD_sum/n = {krd_sum:.3f}")
        ax2.axhline(mod_dur / len(tenors), color="tomato", lw=1.5, ls=":",
                    label=f"ModDur = {mod_dur:.3f}")

    ax2.set_xlabel("Тенор (лет)")
    ax2.set_ylabel("KRD портфеля (лет)")
    ax2.set_title("KRD профиль портфеля по тенорам")
    ax2.legend(fontsize=9)
    ax2.grid(True, alpha=0.3, axis="y")

    plt.tight_layout()
    plt.savefig("krd.png", dpi=150, bbox_inches="tight")
    plt.close()
    print("Сохранён krd.png")

CBR_KEY_RATE = ###

CBR_SCENARIOS = {
    "−100 bps (агрессивное смягчение)": -100,
    "−50 bps (базовый)":                 -50,
    "−25 bps (осторожное)":              -25,
    "0 bps (пауза)":                       0,
    "+25 bps (ужесточение)":             +25,
    "+50 bps (жёсткое)":                 +50,
}

def _interp_gcurve(yearyields: dict, t: float) -> float:

    tenors = sorted(yearyields.keys())
    if t <= tenors[0]:  return yearyields[tenors[0]]
    if t >= tenors[-1]: return yearyields[tenors[-1]]
    for i in range(len(tenors) - 1):
        t0, t1 = tenors[i], tenors[i + 1]
        if t0 <= t <= t1:
            w = (t - t0) / (t1 - t0)
            return yearyields[t0] * (1 - w) + yearyields[t1] * w
    return yearyields[tenors[-1]]

def _bump_gcurve(yearyields: dict,
                 shift_bps: float,
                 transmission: dict | None = None) -> dict:

    if transmission is None:
        transmission = {0.25: 1.00, 0.5: 1.00, 0.75: 1.00, 1.0: 1.00,
                        2.0:  0.65, 3.0:  0.65, 5.0:  0.65,
                        7.0:  0.25, 10.0: 0.25, 15.0: 0.20, 20.0: 0.15}

    s   = shift_bps / 100.0  # в %
    out = {}
    for t, r in yearyields.items():

        keys   = sorted(transmission.keys())
        t_near = min(keys, key=lambda k: abs(k - t))
        coeff  = transmission[t_near]
        out[t] = round(r + s * coeff, 6)
    return out

def compute_cbr_scenarios(zcyc: dict,
                           positions: list[dict],
                           analytics: pd.DataFrame,
                           cf_map: dict[str, list[dict]],
                           scenarios: dict = None,
                           key_rate: float = CBR_KEY_RATE,
                           tenors_out: list[float] | None = None) -> dict:
    from math import exp
    from datetime import date as _date

    if scenarios is None:
        scenarios = CBR_SCENARIOS
    if tenors_out is None:
        tenors_out = sorted(zcyc["yearyields"].keys())

    today      = _date.today()
    yy_base    = zcyc["yearyields"]
    ann_df     = analytics.set_index("SECID")

    def pv_portfolio(yy: dict) -> tuple[float, dict]:
        pv_total    = 0.0
        pv_by_secid = {}
        for pos in positions:
            secid = pos["secid"]
            pv    = 0.0
            for cf in cf_map.get(secid, []):
                try:
                    t = (pd.Timestamp(cf["date"]).date() - today).days / 365.25
                except Exception:
                    continue
                if t <= 0:
                    continue
                r   = _interp_gcurve(yy, t) / 100.0
                pv += cf["amount"] * exp(-r * t)
            pv_by_secid[secid] = pv
            pv_total += pv
        return pv_total, pv_by_secid

    pv_base, pv_base_by_secid = pv_portfolio(yy_base)
    results = {}

    for label, shift_bps in scenarios.items():
        yy_new   = _bump_gcurve(yy_base, shift_bps)
        pv_new, pv_new_by_secid = pv_portfolio(yy_new)
        dp     = pv_new - pv_base
        dp_pct = dp / pv_base * 100 if pv_base else 0

        rows = []
        for pos in positions:
            secid = pos["secid"]
            if secid not in ann_df.index:
                continue
            pv_b = pv_base_by_secid.get(secid, 0)
            pv_s = pv_new_by_secid.get(secid, 0)
            rows.append({
                "SECID":       secid,
                "Name":        str(ann_df.loc[secid].get("Name", secid)),
                "PV_base":     round(pv_b, 2),
                "PV_scenario": round(pv_s, 2),
                "dP":          round(pv_s - pv_b, 2),
                "dP_%":        round((pv_s - pv_b) / pv_b * 100
                                     if pv_b else 0, 4),
            })

        results[label] = {
            "shift_bps": shift_bps,
            "key_rate":  round(key_rate + shift_bps / 100, 4),
            "curve":     yy_new,
            "pv_base":   round(pv_base, 2),
            "pv_new":    round(pv_new, 2),
            "dP":        round(dp, 2),
            "dP_%":      round(dp_pct, 4),
            "positions": pd.DataFrame(rows),
        }

    return results

def plot_cbr_scenarios(cbr_results: dict,
                       zcyc: dict,
                       tenors_out: list[float] | None = None) -> None:
    import matplotlib.pyplot as plt
    import matplotlib.ticker as mticker
    import numpy as np

    if tenors_out is None:
        tenors_out = [0.25, 0.5, 1, 2, 3, 5, 7, 10, 15, 20]

    fig, axes = plt.subplots(1, 2, figsize=(18, 7))
    fig.suptitle(f"Сценарии ЦБ — {date.today()} "
                 f"(текущая ставка {CBR_KEY_RATE}%)",
                 fontsize=13, fontweight="bold")

    ax = axes[0]
    base_yy = zcyc["yearyields"]
    base_t  = sorted(base_yy.keys())
    ax.plot(base_t, [base_yy[t] for t in base_t],
            lw=2.5, color="black", ls="--", label="Базовая G-Curve", zorder=5)

    cmap   = plt.cm.get_cmap("RdYlGn", len(cbr_results))
    labels_sorted = sorted(cbr_results.keys(),
                           key=lambda x: cbr_results[x]["shift_bps"])
    for i, label in enumerate(labels_sorted):
        res  = cbr_results[label]
        t_s  = sorted(res["curve"].keys())
        r_s  = [res["curve"][t] for t in t_s]
        shift = res["shift_bps"]
        color = cmap(i)
        ax.plot(t_s, r_s, lw=1.8, color=color, alpha=0.9,
                label=f"{label} → {res['key_rate']}%")
        # аннотация на 5Y
        if 5 in res["curve"]:
            ax.annotate(f"{res['curve'][5]:.2f}%",
                        xy=(5, res["curve"][5]),
                        fontsize=7, color=color,
                        xytext=(3, 0), textcoords="offset points")

    ax.set_xlabel("Тенор (лет)")
    ax.set_ylabel("Доходность, %")
    ax.set_title("G-Curve по сценариям ЦБ")
    ax.legend(fontsize=7)
    ax.grid(True, alpha=0.3)

    ax2 = axes[1]
    shifts = [cbr_results[l]["shift_bps"] for l in labels_sorted]
    dps    = [cbr_results[l]["dP"]        for l in labels_sorted]
    dpps   = [cbr_results[l]["dP_%"]      for l in labels_sorted]
    colors = [cmap(i) for i in range(len(labels_sorted))]
    xlbls  = [f"{'+' if s > 0 else ''}{s}" for s in shifts]

    bars = ax2.bar(xlbls, dps, color=colors, alpha=0.9, width=0.6)
    ax2.axhline(0, color="black", lw=0.8, ls="--")

    for bar, val, pct in zip(bars, dps, dpps):
        ax2.text(bar.get_x() + bar.get_width() / 2,
                 val + (max(abs(v) for v in dps) * 0.02 * (1 if val >= 0 else -1)),
                 f"{val/1e9:.2f}B\n({pct:+.2f}%)",
                 ha="center", va="bottom" if val >= 0 else "top",
                 fontsize=8, fontweight="bold")

    ax2.yaxis.set_major_formatter(
        mticker.FuncFormatter(lambda v, _: f"{v/1e9:.1f}B" if abs(v) >= 1e9
                              else f"{v/1e6:.0f}M"))
    ax2.set_xlabel("Решение ЦБ (bps к ключевой ставке)")
    ax2.set_ylabel("ΔP портфеля (руб.)")
    ax2.set_title("P&L портфеля при решениях ЦБ")
    ax2.grid(True, alpha=0.3, axis="y")

    plt.tight_layout()
    plt.savefig("cbr_scenarios.png", dpi=150, bbox_inches="tight")
    plt.close()
    print("Сохранён cbr_scenarios.png")

if __name__ == "__main__":

    TICKERS_RUB: list[str] = [

    ]

    TICKERS_FX: list[str] = [

    ]

    TICKERS    = TICKERS_RUB
    ENTRY_DATE = "2026-06-26"
    SHIFTS     = [-300, -200, -100, +100, +200, +300]

    print("Резолвинг ISIN → SECID...")
    isin_to_secid: dict[str, str] = {}
    for isin in TICKERS:
        secid = resolve_secid(isin)
        isin_to_secid[isin] = secid
        print(f"  {isin} → {secid}")

    QTY: dict[str, int] = {

    }

    print(f"\nЗагрузка цен входа на {ENTRY_DATE}...")
    positions = []
    for isin, secid in isin_to_secid.items():
        if secid is None:
            continue
        cost = get_entry_price(secid, ENTRY_DATE)
        qty  = QTY.get(isin, 100)
        positions.append({"secid": secid, "qty": qty, "cost_price": cost})
        print(f"  {isin} / {secid}: qty={qty}, cost={cost} руб.")

    try:
        zcyc = get_moex_zcyc()
    except Exception as e:
        zcyc = None
        print(f"[WARN] G-Curve MOEX недоступна: {e}")

    pf        = OFZPortfolio(positions)
    analytics = pf.compute(zcyc=zcyc)
    print("\n=== Позиционная аналитика ===")
    print(analytics.to_string(index=False))

    coupon_candidates = [c for c in analytics.columns
                         if any(x in c.lower() for x in
                                ["coupon", "face", "rate", "nominal"])]
    print(f"\n[DEBUG] Колонки с купоном/номиналом: {coupon_candidates}")
    if coupon_candidates:
        print(analytics[["SECID"] + coupon_candidates].to_string(index=False))

    COL_MAP = {
        # купон — ищем первое совпадение
        "CouponPct":  ["CouponPct", "couponpercent", "COUPONPERCENT",
                       "CouponRate", "coupon_rate", "coupon_pct"],
        # номинал
        "Face":       ["Face", "FACEVALUE", "facevalue",
                       "face_value", "FaceValue", "nominal"],
    }
    for target, variants in COL_MAP.items():
        if target not in analytics.columns:
            for v in variants:
                if v in analytics.columns:
                    analytics[target] = analytics[v]
                    print(f"  [NORM] {v} → {target}")
                    break
            else:
                print(f"  [WARN] {target} не найден в analytics — "
                      f"попытка загрузить с ISS...")
                vals = {}
                for _, row in analytics.iterrows():
                    secid = row["SECID"]
                    try:
                        r = requests.get(
                            f"{BASE_ISS}/securities/{secid}.json",
                            params={"iss.meta": "off"}, timeout=10
                        ).json()
                        for d in r.get("description", {}).get("data", []):
                            if target == "CouponPct" and \
                               d[0] in ("COUPONPERCENT",) and d[2]:
                                vals[secid] = float(d[2]); break
                            if target == "Face" and \
                               d[0] in ("FACEVALUE",) and d[2]:
                                vals[secid] = float(d[2]); break
                    except Exception:
                        pass
                if vals:
                    analytics[target] = analytics["SECID"].map(vals).fillna(
                        1000 if target == "Face" else 0)
                    print(f"  [ISS]  {target} загружен для "
                          f"{len(vals)}/{len(analytics)} позиций")
                else:
                    analytics[target] = 1000 if target == "Face" else 0

    smry = pf.summary()
    if smry:
        print("\n=== Агрегат ===")
        for k, v in smry.items():
            print(f"  {k}: {v}")

    stress_data = {}
    for shift in SHIFTS:
        stress = pf.scenario_shift(shift)
        stress_data[shift] = stress
        if not stress.empty:
            sign = "+" if shift > 0 else ""
            print(f"\n=== Сценарий: {sign}{shift} bps ===")
            print(stress.to_string(index=False))
            print(f"  Итого ΔP         (руб.): {stress['dP'].sum():.2f}")
            mv_stressed       = stress["MktValue"] + stress["dP"]
            w_stressed        = mv_stressed / mv_stressed.sum()
            port_mac_base     = (stress["MacDur_years"]    * w_stressed).sum()
            port_mac_stressed = (stress["MacDur_stressed"] * w_stressed).sum()
            port_mod_base     = (stress["ModDur_years"]    * w_stressed).sum()
            port_mod_stressed = (stress["ModDur_stressed"] * w_stressed).sum()
            print(f"  PortMacDur до    (лет):  {port_mac_base:.4f}")
            print(f"  PortMacDur после (лет):  {port_mac_stressed:.4f}")
            print(f"  ΔPortMacDur      (лет):  {port_mac_stressed - port_mac_base:.4f}")
            print(f"  PortModDur до    (лет):  {port_mod_base:.4f}")
            print(f"  PortModDur после (лет):  {port_mod_stressed:.4f}")
            print(f"  ΔPortModDur      (лет):  {port_mod_stressed - port_mod_base:.4f}")
            print(f"  PortDV01 до      (руб.): {stress['DV01'].sum():.2f}")
            print(f"  PortDV01 после   (руб.): {stress['DV01_stressed'].sum():.2f}")
            print(f"  ΔPortDV01        (руб.): {stress['dDV01'].sum():.2f}")

    attr = pf.pnl_attribution()
    if not attr.empty:
        print("\n=== P&L Attribution ===")
        print(attr.to_string(index=False))

    import os
    import matplotlib.pyplot as plt
    import matplotlib.gridspec as gridspec
    from scipy.interpolate import CubicSpline
    import matplotlib.ticker as mticker

    secid_labels = analytics["SECID"].str[-7:].tolist()
    colors_shift = {
        -300: "#0d3b6e", -200: "#1a6fad", -100: "#5aaee0",
        +100: "#e07b5a", +200: "#ad2a1a", +300: "#5c0f0f",
    }

    fig = plt.figure(figsize=(20, 22))
    gs  = gridspec.GridSpec(4, 2, figure=fig, hspace=0.45, wspace=0.35)

    ax1 = fig.add_subplot(gs[0, :])
    if zcyc is not None:
        yy     = zcyc["yearyields"]
        tenors = sorted(yy.keys())
        rates  = [yy[t] for t in tenors]
        cs     = CubicSpline(tenors, rates)
        t_grid = np.linspace(min(tenors), max(tenors), 300)
        ax1.plot(t_grid, cs(t_grid), lw=2, color="steelblue", label="G-Curve MOEX")
        ax1.scatter(tenors, rates, s=30, color="steelblue", zorder=4)
    ax1.scatter(analytics["ModDur_years"], analytics["YTM_%"],
                s=90, color="tomato", zorder=5, label="Позиции")
    for _, r in analytics.iterrows():
        ax1.annotate(r["SECID"][-7:], (r["ModDur_years"], r["YTM_%"]),
                     textcoords="offset points", xytext=(6, 5), fontsize=8)
    ax1.set_xlabel("Modified Duration, лет")
    ax1.set_ylabel("Доходность, %")
    title_date = zcyc["tradedate"] if zcyc else str(date.today())
    ax1.set_title(f"G-Curve MOEX + позиции портфеля — {title_date}")
    ax1.legend(); ax1.grid(True, alpha=0.3)

    ax2    = fig.add_subplot(gs[1, :])
    n      = len(secid_labels)
    x      = np.arange(n)
    BAR_W2 = 0.10; BAR_STEP2 = 0.135
    for i, shift in enumerate(SHIFTS):
        st = stress_data.get(shift)
        if st is None or st.empty:
            continue
        ax2.bar(x + (i - 2.5) * BAR_STEP2, st["dP"], BAR_W2,
                label=f"{'+' if shift>0 else ''}{shift} bps",
                color=colors_shift[shift], alpha=0.85)
    ax2.axhline(0, color="black", lw=0.8, ls="--")
    ax2.set_xticks(x); ax2.set_xticklabels(secid_labels, rotation=15)
    ax2.set_ylabel("ΔP (руб.)"); ax2.set_title("Изменение стоимости позиций по сценариям")
    ax2.legend(); ax2.grid(True, alpha=0.3, axis="y")

    BAR_W = 0.09; BAR_STEP = 0.12; BASE_W = 0.10
    BASE_OFF = -(2.5 * BAR_STEP + 0.18)

    ax3         = fig.add_subplot(gs[2, 0])
    base_moddur = analytics.set_index("SECID")["ModDur_years"]
    ax3.bar(x + BASE_OFF, base_moddur.values, BASE_W,
            label="Base", color="steelblue", alpha=0.9)
    for i, shift in enumerate(SHIFTS):
        st = stress_data.get(shift)
        if st is None or st.empty: continue
        ax3.bar(x + (i - 2.5) * BAR_STEP, st["ModDur_stressed"], BAR_W,
                label=f"{'+' if shift>0 else ''}{shift} bps",
                color=colors_shift[shift], alpha=0.75)
    ax3.set_xticks(x); ax3.set_xticklabels(secid_labels, rotation=15, fontsize=8)
    ax3.set_ylabel("Modified Duration, лет"); ax3.set_title("ModDur: базовый vs стресс")
    ax3.legend(fontsize=7); ax3.grid(True, alpha=0.3, axis="y")

    ax4         = fig.add_subplot(gs[2, 1])
    base_macdur = analytics.set_index("SECID")["MacDur_years"]
    ax4.bar(x + BASE_OFF, base_macdur.values, BASE_W,
            label="Base", color="steelblue", alpha=0.9)
    for i, shift in enumerate(SHIFTS):
        st = stress_data.get(shift)
        if st is None or st.empty: continue
        ax4.bar(x + (i - 2.5) * BAR_STEP, st["MacDur_stressed"], BAR_W,
                label=f"{'+' if shift>0 else ''}{shift} bps",
                color=colors_shift[shift], alpha=0.75)
    ax4.set_xticks(x); ax4.set_xticklabels(secid_labels, rotation=15, fontsize=8)
    ax4.set_ylabel("Macaulay Duration, лет"); ax4.set_title("MacDur: базовый vs стресс")
    ax4.legend(fontsize=7); ax4.grid(True, alpha=0.3, axis="y")

    ax5 = fig.add_subplot(gs[3, 0])
    ax5.bar(x + BASE_OFF, analytics["DV01"], BASE_W,
            label="Base", color="steelblue", alpha=0.9)
    for i, shift in enumerate(SHIFTS):
        st = stress_data.get(shift)
        if st is None or st.empty: continue
        ax5.bar(x + (i - 2.5) * BAR_STEP, st["DV01_stressed"], BAR_W,
                label=f"{'+' if shift>0 else ''}{shift} bps",
                color=colors_shift[shift], alpha=0.75)
    ax5.set_xticks(x); ax5.set_xticklabels(secid_labels, rotation=15, fontsize=8)
    ax5.set_ylabel("DV01 (руб. / 1 bp)"); ax5.set_title("DV01: базовый vs стресс")
    ax5.legend(fontsize=7); ax5.grid(True, alpha=0.3, axis="y")

    ax6       = fig.add_subplot(gs[3, 1])
    total_dps = [stress_data[s]["dP"].sum() for s in SHIFTS if s in stress_data]
    bar_colors = [colors_shift[s] for s in SHIFTS if s in stress_data]
    shift_lbls = [f"{'+' if s>0 else ''}{s}" for s in SHIFTS if s in stress_data]
    bars = ax6.bar(shift_lbls, total_dps, color=bar_colors, alpha=0.9, width=0.5)
    ax6.axhline(0, color="black", lw=0.8, ls="--")
    for bar, val in zip(bars, total_dps):
        ax6.text(bar.get_x() + bar.get_width() / 2,
                 val + (500 if val >= 0 else -1500),
                 f"{val:,.0f} руб.", ha="center", fontsize=9, fontweight="bold")
    ax6.set_xlabel("Сдвиг кривой, bps")
    ax6.set_ylabel("ΔP портфеля (руб.)")
    ax6.set_title("Суммарный P&L портфеля по сценариям")
    ax6.grid(True, alpha=0.3, axis="y")

    plt.suptitle(f"OFZ Portfolio Analytics — {date.today()}", fontsize=14,
                 fontweight="bold", y=1.01)
    plt.savefig("portfolio_analytics.png", dpi=150, bbox_inches="tight")
    plt.close()
    print("\nСохранён portfolio_analytics.png")

    print("\nСтроим Cash Flow Ladder...")
    ladder_wide = pd.DataFrame()
    ladder_agg  = pd.DataFrame()
    try:
        ladder_wide, ladder_agg = build_cf_ladder(analytics, positions, freq="ME")
    except Exception as e:
        print(f"[WARN] CF Ladder не построен: {e}")

    if not ladder_agg.empty:
        print("\n=== CF Ladder (агрегат, первые 12 периодов) ===")
        print(ladder_agg[["period_str", "coupon",
                           "amortization", "total"]].head(12).to_string(index=False))
        plot_cf_ladder(ladder_agg, ladder_wide, freq_label="месяц")
    else:
        print("[WARN] CF Ladder пустой — cf_ladder.png не создан")

    print("\nBreak-Even Analysis...")
    for horizon in [90, 180, 365]:
        be_df = breakeven_analysis(analytics, positions, horizon_days=horizon)
        if not be_df.empty:
            print(f"\n=== Break-Even (горизонт {horizon} дн.) ===")
            print(be_df.to_string(index=False))

    be_df = breakeven_analysis(analytics, positions, horizon_days=365)
    if not be_df.empty:
        plot_breakeven(be_df, horizon_days=365)
    else:
        print("[WARN] Break-Even не рассчитан — breakeven.png не создан")
        be_df = pd.DataFrame()

    krd_df = pd.DataFrame()
    if zcyc is not None:
        print("\nKRD Analysis...")
        cf_map: dict[str, list[dict]] = {}
        ann_df = analytics.set_index("SECID")

        for pos in positions:
            secid = pos["secid"]
            qty   = pos.get("qty", 1)
            face  = float(ann_df.loc[secid, "Face"]) \
                    if secid in ann_df.index else 1000.0
            try:
                raw_cf = get_cashflows(secid, qty, face)

                if isinstance(raw_cf, pd.DataFrame):
                    if raw_cf.empty:
                        print(f"  [WARN] {secid}: пустой CF DataFrame")
                        continue

                    if not cf_map:
                        print(f"  [DEBUG] CF columns: {raw_cf.columns.tolist()}")
                        print(raw_cf.head(2).to_string())

                    date_col = next((c for c in raw_cf.columns if c.lower() in
                                     ("date", "tradedate", "coupondate",
                                      "cf_date", "period", "paydate")), None)
                    amt_col  = next((c for c in raw_cf.columns if c.lower() in
                                     ("amount", "total", "value", "cf",
                                      "cashflow", "payment", "sum",
                                      "amount_rub")), None)

                    if date_col is None or amt_col is None:
                        print(f"  [WARN] {secid}: не найдены колонки "
                              f"date/amount → {raw_cf.columns.tolist()}")
                        continue

                    cf_map[secid] = [
                        {"date": row[date_col], "amount": float(row[amt_col])}
                        for _, row in raw_cf.iterrows()
                        if pd.notna(row[amt_col]) and float(row[amt_col]) > 0
                    ]

                elif isinstance(raw_cf, list) and raw_cf \
                        and isinstance(raw_cf[0], dict):
                    cf_map[secid] = [
                        {"date": cf["date"], "amount": float(cf["amount"])}
                        for cf in raw_cf
                        if cf.get("date") and float(cf.get("amount", 0)) > 0
                    ]

                elif isinstance(raw_cf, list) and raw_cf \
                        and isinstance(raw_cf[0], (list, tuple)):
                    if not cf_map:
                        print(f"  [DEBUG] CF как список списков, "
                              f"первый элемент: {raw_cf[0]}")
                    cf_map[secid] = [
                        {"date": cf[0], "amount": float(cf[-1])}
                        for cf in raw_cf
                        if cf[0] and float(cf[-1]) > 0
                    ]

                else:
                    print(f"  [WARN] {secid}: неизвестный тип "
                          f"get_cashflows → {type(raw_cf)}, "
                          f"первый эл.: {raw_cf[0] if raw_cf else 'пусто'}")
                    continue

                n_cf = len(cf_map.get(secid, []))
                if n_cf == 0:
                    print(f"  [WARN] {secid}: CF распарсен, но 0 событий")
                    cf_map.pop(secid, None)
                else:
                    print(f"  [KRD] {secid}: {n_cf} CF событий")

            except Exception as e:
                print(f"  [WARN] CF для {secid}: {e}")

        if cf_map:
            _test_secid = list(cf_map.keys())[0]
            _test_pos   = next(p for p in positions if p["secid"] == _test_secid)
            _ann        = analytics.set_index("SECID").loc[_test_secid]
            print(f"\n[DEBUG] compute_krd тест на {_test_secid}:")
            print(f"  MktValue  = {_ann.get('MktValue',  'НЕТ')}")
            print(f"  MktPrice  = {_ann.get('MktPrice',  'НЕТ')}")
            print(f"  Face      = {_ann.get('Face',      'НЕТ')}")
            print(f"  ModDur    = {_ann.get('ModDur_years', 'НЕТ')}")
            print(f"  CF[0]     = {cf_map[_test_secid][0]}")
            print(f"  zcyc keys = {list(zcyc.keys())}")
            print(f"  yearyields= {zcyc.get('yearyields', 'НЕТ')}")

            krd_df = compute_krd(analytics, positions, zcyc, cf_map, shift_bps=1.0)
            krd_df = compute_krd(analytics, positions, zcyc, cf_map, shift_bps=1.0)
            if not krd_df.empty:
                print("\n=== Key Rate Duration ===")
                print(krd_df.to_string(index=False))
                plot_krd(krd_df)
            else:
                print("[WARN] KRD не рассчитан — пустой результат")
        else:
            print("[WARN] KRD пропущен: cf_map пустой (см. [DEBUG] выше)")
    else:
        print("[WARN] KRD пропущен: нет G-Curve (zcyc=None)")
    
    cbr_results = {}
    if zcyc is not None and "params" in zcyc and cf_map:
        print("\nСценарии ЦБ...")

        cbr_results = compute_cbr_scenarios(
            zcyc, positions, analytics, cf_map,
            scenarios=CBR_SCENARIOS, key_rate=CBR_KEY_RATE,
        )
        for label, res in sorted(cbr_results.items(),
                                  key=lambda x: x[1]["shift_bps"]):
            print(f"\n  [{'+' if res['shift_bps']>0 else ''}"
                  f"{res['shift_bps']:+d} bps] {label}")
            print(f"    Ключевая ставка: {CBR_KEY_RATE}% → {res['key_rate']}%")
            print(f"    ΔP портфеля:     {res['dP']:>+,.0f} руб. "
                  f"({res['dP_%']:+.4f}%)")
            print(f"    G-Curve 1Y:  {res['curve'].get(1, 0):.4f}%  "
                  f"5Y: {res['curve'].get(5, 0):.4f}%  "
                  f"10Y: {res['curve'].get(10, 0):.4f}%")
        plot_cbr_scenarios(cbr_results, zcyc)
    else:
        print("[WARN] Сценарии ЦБ пропущены: нет zcyc['params'] или cf_map")

    import os

    XLSX_PATH = "ofz_portfolio_analytics.xlsx"
    wb        = openpyxl.Workbook()

    fx_analytics = pd.DataFrame()
    if TICKERS_FX:
        fx_isin_to_secid: dict[str, str] = {}
        for isin in TICKERS_FX:
            s = resolve_secid(isin)
            if s:
                fx_isin_to_secid[isin] = s
                print(f"  [FX] {isin} → {s}")
        fx_positions = [
            {"secid": sid, "qty": QTY.get(isin, 0), "cost_price": None}
            for isin, sid in fx_isin_to_secid.items()
        ]
        if fx_positions:
            fx_pf        = OFZPortfolio(fx_positions)
            fx_analytics = fx_pf.compute(zcyc=None)

    ws1       = wb.active
    ws1.title = "Аналитика"
    n_cols    = len(analytics.columns)

    title_cell           = ws1["A1"]
    title_cell.value     = f"OFZ Portfolio Analytics — {date.today()}"
    title_cell.font      = Font(bold=True, size=13, color="1F4E79")
    title_cell.alignment = CENTER
    ws1.merge_cells(f"A1:{chr(64 + n_cols)}1")
    ws1.row_dimensions[1].height = 22

    next_row = write_df(ws1, analytics, start_row=3)

    if not fx_analytics.empty:
        next_row += 1
        hdr           = ws1.cell(row=next_row, column=1,
                                 value="⚠️  Валютные позиции — справочно, "
                                       "не включены в расчёты")
        hdr.font      = Font(bold=True, size=10, color="7F6000")
        hdr.fill      = FX_NOTE
        hdr.alignment = LEFT
        ws1.merge_cells(f"A{next_row}:{chr(64 + n_cols)}{next_row}")
        ws1.row_dimensions[next_row].height = 20
        next_row += 1

        for c_idx, col_name in enumerate(fx_analytics.columns, 1):
            cell           = ws1.cell(row=next_row, column=c_idx, value=col_name)
            cell.fill      = PatternFill("solid", fgColor="FFE699")
            cell.font      = Font(bold=True, color="7F6000", size=10)
            cell.alignment = CENTER
            cell.border    = BORDER
        next_row += 1

        for _, frow in fx_analytics.iterrows():
            for c_idx, val in enumerate(frow, 1):
                cell           = ws1.cell(row=next_row, column=c_idx, value=val)
                cell.font      = FX_FONT
                cell.fill      = FX_FILL
                cell.border    = BORDER
                cell.alignment = CENTER
            next_row += 1

        note           = ws1.cell(row=next_row, column=1,
                                  value="Метрики валютных бумаг указаны в валюте "
                                        "номинала и несопоставимы с рублёвыми позициями выше.")
        note.font      = Font(italic=True, size=9, color="7F6000")
        note.alignment = LEFT
        ws1.merge_cells(f"A{next_row}:{chr(64 + n_cols)}{next_row}")
        ws1.row_dimensions[next_row].height = 20

    autofit(ws1)

    ws2             = wb.create_sheet("Агрегат")
    ws2["A1"].value = "Метрика"
    ws2["B1"].value = "Значение"
    style_header(ws2, 1)
    for i, (k, v) in enumerate(smry.items(), 2):
        ws2.cell(row=i, column=1, value=k)
        ws2.cell(row=i, column=2, value=v)
    style_rows(ws2, 2, 1 + len(smry))
    note_row = 2 + len(smry) + 1
    ws2.cell(row=note_row, column=1,
             value="Агрегат рассчитан только по рублёвым позициям. "
                   "Валютные бумаги вынесены на лист «Аналитика» справочно.")
    ws2.cell(row=note_row, column=1).font      = Font(italic=True, size=9, color="595959")
    ws2.cell(row=note_row, column=1).alignment = LEFT
    ws2.merge_cells(f"A{note_row}:B{note_row}")
    autofit(ws2)

    ws3     = wb.create_sheet("Стресс-сценарии")
    cur_row = 1
    ws3.cell(row=cur_row, column=1,
             value="Стресс-сценарии рассчитаны только по рублёвым позициям.")
    ws3.cell(row=cur_row, column=1).font      = Font(italic=True, size=9, color="595959")
    ws3.cell(row=cur_row, column=1).alignment = LEFT
    cur_row += 2

    for shift in SHIFTS:
        st = stress_data.get(shift)
        if st is None or st.empty:
            continue
        sign  = "+" if shift > 0 else ""
        label = f"Сценарий: {sign}{shift} bps"
        ws3.cell(row=cur_row, column=1, value=label)
        ws3.cell(row=cur_row, column=1).font      = Font(bold=True, size=11, color="FFFFFF")
        ws3.cell(row=cur_row, column=1).fill      = SUB_FILL
        ws3.cell(row=cur_row, column=1).alignment = CENTER
        ws3.merge_cells(f"A{cur_row}:{chr(64 + len(st.columns))}{cur_row}")
        ws3.row_dimensions[cur_row].height = 20
        cur_row += 1
        cur_row = write_df(ws3, st, start_row=cur_row)
        ws3.cell(row=cur_row, column=1, value="Итого ΔP (руб.)")
        ws3.cell(row=cur_row, column=1).font = Font(bold=True)
        dp_col = list(st.columns).index("dP") + 1
        ws3.cell(row=cur_row, column=dp_col, value=round(st["dP"].sum(), 2))
        ws3.cell(row=cur_row, column=dp_col).font = Font(bold=True, color="1F4E79")
        cur_row += 2

    autofit(ws3)

    ws4 = wb.create_sheet("PnL Attribution")
    write_df(ws4, attr, start_row=1)
    autofit(ws4)

    if os.path.exists("portfolio_analytics.png"):
        ws5            = wb.create_sheet("График")
        img5           = XLImage("portfolio_analytics.png")
        orig_w, orig_h = img5.width, img5.height
        img5.width     = 900
        img5.height    = int(900 * orig_h / orig_w)
        img5.anchor    = "A1"
        ws5.add_image(img5)

    if not ladder_agg.empty:
        ws_agg          = wb.create_sheet("CF Агрегат")
        out_agg         = ladder_agg[["period_str", "coupon",
                                      "amortization", "total"]].copy()
        out_agg.columns = ["Период", "Купоны (руб.)",
                           "Погашение (руб.)", "Итого (руб.)"]
        write_df(ws_agg, out_agg, start_row=1)
        tot_row = 2 + len(out_agg)
        ws_agg.cell(row=tot_row, column=1, value="ИТОГО").font = Font(bold=True)
        col_map = {"Купоны (руб.)":     "coupon",
                   "Погашение (руб.)":  "amortization",
                   "Итого (руб.)":      "total"}
        for col_i, col_name in enumerate(col_map.keys(), 2):
            cell      = ws_agg.cell(row=tot_row, column=col_i,
                                    value=round(ladder_agg[col_map[col_name]].sum(), 2))
            cell.font = Font(bold=True, color="1F4E79")
            cell.fill = PatternFill("solid", fgColor="DEEAF1")
        autofit(ws_agg)

        ws_wide  = wb.create_sheet("CF По бумагам")
        out_wide = ladder_wide.drop(columns=["period"]).copy()
        out_wide = out_wide.rename(columns={"period_str": "Период"})
        write_df(ws_wide, out_wide, start_row=1)
        autofit(ws_wide)

        if os.path.exists("cf_ladder.png"):
            ws_img         = wb.create_sheet("CF График")
            img_cf         = XLImage("cf_ladder.png")
            orig_w, orig_h = img_cf.width, img_cf.height
            img_cf.width   = 900
            img_cf.height  = int(900 * orig_h / orig_w)
            img_cf.anchor  = "A1"
            ws_img.add_image(img_cf)

    if not be_df.empty:
        ws_be = wb.create_sheet("Break-Even")
        write_df(ws_be, be_df, start_row=1)
        last_row = 1 + len(be_df) + 1
        for cell in ws_be[last_row]:
            cell.font = Font(bold=True, size=10, color="1F4E79")
            cell.fill = PatternFill("solid", fgColor="DEEAF1")
        autofit(ws_be)

        if os.path.exists("breakeven.png"):
            ws_bepng       = wb.create_sheet("BE График")
            img_be         = XLImage("breakeven.png")
            orig_w, orig_h = img_be.width, img_be.height
            img_be.width   = 900
            img_be.height  = int(900 * orig_h / orig_w)
            img_be.anchor  = "A1"
            ws_bepng.add_image(img_be)

    if not krd_df.empty:
        ws_krd = wb.create_sheet("KRD")
        ws_krd["A1"].value     = ("Key Rate Duration — чувствительность "
                                   "к сдвигу каждого тенора на 1 bps")
        ws_krd["A1"].font      = Font(italic=True, size=9, color="595959")
        ws_krd["A1"].alignment = LEFT
        ws_krd.merge_cells(f"A1:{chr(64 + len(krd_df.columns))}1")
        write_df(ws_krd, krd_df, start_row=2)
        last_row = 2 + len(krd_df) + 1
        for cell in ws_krd[last_row]:
            cell.font = Font(bold=True, size=10, color="1F4E79")
            cell.fill = PatternFill("solid", fgColor="DEEAF1")
        autofit(ws_krd)

        if os.path.exists("krd.png"):
            ws_krd_img     = wb.create_sheet("KRD График")
            img_krd        = XLImage("krd.png")
            orig_w, orig_h = img_krd.width, img_krd.height
            img_krd.width  = 900
            img_krd.height = int(900 * orig_h / orig_w)
            img_krd.anchor = "A1"
            ws_krd_img.add_image(img_krd)

    if cbr_results:
        ws_cbr      = wb.create_sheet("Сценарии ЦБ")
        summary_rows = []
        for label, res in sorted(cbr_results.items(),
                                  key=lambda x: x[1]["shift_bps"]):
            row = {
                "Сценарий":              label,
                "Ставка ЦБ (%)":         res["key_rate"],
                "Δ Ключ. ставка (bps)":  res["shift_bps"],
                "PV базовый (руб.)":     res["pv_base"],
                "PV сценарий (руб.)":    res["pv_new"],
                "ΔP (руб.)":             res["dP"],
                "ΔP (%)":               res["dP_%"],
            }
            for t in [1, 3, 5, 10]:
                row[f"G-Curve {t}Y (%)"] = res["curve"].get(t, "")
            summary_rows.append(row)

        write_df(ws_cbr, pd.DataFrame(summary_rows), start_row=1)
        autofit(ws_cbr)

        ws_cbrd = wb.create_sheet("CBR Детализация")
        cur_row = 1
        for label, res in sorted(cbr_results.items(),
                                  key=lambda x: x[1]["shift_bps"]):
            if res["positions"].empty:
                continue
            n_cols_cbr = len(res["positions"].columns)
            ws_cbrd.cell(row=cur_row, column=1, value=label)
            ws_cbrd.cell(row=cur_row, column=1).font      = Font(bold=True, size=11,
                                                                  color="FFFFFF")
            ws_cbrd.cell(row=cur_row, column=1).fill      = SUB_FILL
            ws_cbrd.cell(row=cur_row, column=1).alignment = CENTER
            ws_cbrd.merge_cells(
                f"A{cur_row}:{chr(64 + n_cols_cbr)}{cur_row}")
            ws_cbrd.row_dimensions[cur_row].height = 20
            cur_row += 1
            cur_row  = write_df(ws_cbrd, res["positions"], start_row=cur_row)
            cur_row += 1
        autofit(ws_cbrd)

        if os.path.exists("cbr_scenarios.png"):
            ws_cbr_img     = wb.create_sheet("CBR График")
            img_cbr        = XLImage("cbr_scenarios.png")
            orig_w, orig_h = img_cbr.width, img_cbr.height
            img_cbr.width  = 900
            img_cbr.height = int(900 * orig_h / orig_w)
            img_cbr.anchor = "A1"
            ws_cbr_img.add_image(img_cbr)

    wb.save(XLSX_PATH)
    print(f"Сохранён {XLSX_PATH}")
